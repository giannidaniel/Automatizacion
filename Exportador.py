import pandas as pd
from openpyxl import load_workbook

class Exportador():
    
    try:
        matriz = pd.read_excel(io="Matriz.xlsx",sheet_name="Worksheet",index_col=0)
    except:
        print("Falta archivo excel con la matriz")
        
    df=pd.DataFrame(index=matriz.index,columns=["Comentario","Resultado"])
    
    
    def escribir_resultado(self,control,comentario,resultado):
        self.df["Comentario"][control]=comentario
        self.df["Resultado"][control]=resultado
        

    def guardar_salida(self):
        wb=load_workbook("Matriz.xlsx")
        ws=wb.active
        #self.df[]
        for fila in self.df.index:
            ws.cell(row=self.matriz.index.get_loc(fila)+2,column=self.matriz.columns.get_loc("Comentario")+2).value=self.df["Comentario"][fila]
            ws.cell(row=self.matriz.index.get_loc(fila)+2,column=self.matriz.columns.get_loc("Resultado")+2).value=self.df["Resultado"][fila] 
        wb.save("prueba.xlsx")
    
